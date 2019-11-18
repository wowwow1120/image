# /!/usr/bin/python
# -*-coding: utf-8 -*-

import os
import json
import PIL.Image
import PIL.ImageTk
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Color
from operator import itemgetter
from enum import Enum
import numpy as np
from itertools import chain
from collections import defaultdict
import traceback

camera_id = 'camera_id'
date = 'taken_date'
time = 'taken_time'
plate_num = 'plt_num'
camera_uri = 'uri'
veh_uri = 'veh_uri'
plate_uri = 'plt_uri'
veh_type = 'type'  # 0:other, 1:car, 2:van, 3:bus, 4:truck, 5:plate, 6:window
veh_color = 'color'  # 0:unknown, 1:white, 2:silver, 3:gray, 4:black, 5:red, 6:darkblue, 7:blue, 8:yellow, 9:green, 10:brown, 11:pink, 12:purple, 13:darkgray, 14:cyan
veh_dir = 'veh_dir'  # 0:unknown, 1:entry, 2:exit
confidence = 'plt_conf'
lane_info = 'lane_info'  # 0~9
veh_side = 'veh_side'  # 0:unknown, 1:front, 2:back


class VehicleType(Enum):  # 차량의 전면/후면
    other = 0
    car = 1
    van = 2
    bus = 3
    truck = 4
    plate = 5
    window = 6


class VehicleColor(Enum):  # 차량의 전면/후면
    unknown = 0
    white = 1
    silver = 2
    gray = 3
    black = 4
    red = 5
    darkblue = 6
    blue = 7
    yellow = 8
    green = 9
    brown = 10
    pink = 11
    purple = 12
    darkgray = 13
    cyan = 14


class VehicleSide(Enum):  # 차량의 전면/후면
    unknown = 0
    front = 1
    back = 2


class VehicleDirection(Enum):  # 진입/진출
    unknown = 0
    entry = 1
    exit = 2


def window_position(frame, width, height):
    screen_width = frame.winfo_screenwidth()
    screen_height = frame.winfo_screenheight()
    x = (screen_width/2) - (width/2) + (screen_width/6)
    y = (screen_height/2) - (height/2)
    frame.geometry('%dx%d+%d+%d' % (width, height, x, y))


def get_accuracy(result_dict, cor="O", incor="X", quest="q", impossible="XX", special_car="*"):

    cor_num = 0
    incor_num = 0
    quest_num = 0
    impossible_num = 0
    special_num = 0

    acc_list = np.array(list(result_dict.values())).T.tolist()[-1]
    for result in acc_list:
        if result == cor:
            cor_num += 1
        if result == incor:
            incor_num += 1
        if result == quest:
            quest_num += 1
        if result == impossible:
            impossible_num += 1
        if result == special_car:
            special_num += 1

    accuracy = '{:4.2f}%'.format((cor_num / (cor_num + incor_num + quest_num + impossible_num + special_num)) * 100)
    return cor_num, incor_num, quest_num, impossible_num, special_num, accuracy


def merge_dicts(fir_dict, sec_dict):
    result_dict = defaultdict(list)
    for k, v in chain(fir_dict.items(), sec_dict.items()):
        result_dict[k].extend(v)

    return result_dict


class DetectCheck:
    def __init__(self, master):

        self.parent = master
        self.parent.title("검지율 체크 도구")
        self.frame = Frame(self.parent)
        self.frame.pack(fill=BOTH, expand=True)
        self.parent.resizable(width=True, height=True)

        self.Labeling_result = ''
        self.input_dir = ''
        self.json_list = []
        self.img_list = []
        self.cname_list = []
        self.json_info_dict = {'0': ['camera_id', 'taken_date', 'taken_time', 'uri'],
                               '1': ['veh_uri', 'plt_uri', 'plt_num', 'type', 'color', 'veh_dir', 'veh_side',
                                     'lane_info', 'plt_conf']}
        self.camera_info = {}
        self.img = None
        self.camera_img = None
        self.car_img = None
        self.plt_img = None
        self.tkimg = None
        self.answer_value = None
        self.answer_dict = {}  # {self.cur : answer}
        self.cur = 0
        self.plt_type_dict = {}  # {self.cur: [plt_type, plt_uri, detect_result]}
        self.result = {}
        self.result_merged_dict = {}

        try:
            self.det_img = PIL.ImageTk.PhotoImage(PIL.Image.open("detective.gif").resize((170, 170)))
        except FileNotFoundError:
            self.det_img = None

        # GUI
        # input_dir_button
        self.srcDirBtn = Button(self.frame, text="Input folder", command=self.select_input_dir)
        self.srcDirBtn.grid(row=0, column=0, padx=2, pady=2, ipadx=5, ipady=5, sticky=W + E)

        # input_dir_entry
        self.source_dir = StringVar()
        self.entrySrc = Entry(self.frame, textvariable=self.source_dir)
        self.entrySrc.grid(row=0, column=1, padx=2, pady=2, ipadx=5, ipady=5, sticky=W + E)

        # load_button
        self.ldBtn = Button(self.frame, text="Load Dir", command=self.load_all)
        self.ldBtn.grid(row=0, column=2, padx=2, pady=2, ipadx=5, ipady=5, sticky=W + E)

        # export button
        self.expBtn = Button(self.frame, text="Export Excel", command=self.export_button)
        self.expBtn.grid(row=0, column=3, padx=2, pady=2, ipadx=5, ipady=5)

        # main_panel
        self.mainPanel_1 = Canvas(self.frame, width=500, height=500)
        self.mainPanel_1.grid(row=1, column=0, sticky=N)
        self.mainPanel_2 = Canvas(self.frame, width=300, height=300)
        self.mainPanel_2.grid(row=1, column=1, rowspan=2, sticky=N)
        self.mainPanel_3 = Canvas(self.frame, width=300, height=200)
        self.mainPanel_3.grid(row=1, column=1, sticky=S)
        self.mainPanel_4 = Canvas(self.frame, width=250, height=500)
        self.mainPanel_4.grid(row=1, column=2, rowspan=2, columnspan=2, sticky=NE)
        self.mainPanel_5 = Canvas(self.frame, width=250, height=190)
        self.mainPanel_5.grid(row=1, column=2, rowspan=2, columnspan=2, sticky=SE)

        # plate type button
        self.oneBtn = Button(self.frame, text="1(일반)", bg="skyblue", command=self.one_button)
        self.oneBtn.place(x=5, y=445, height=30, width=65)
        self.twoBtn = Button(self.frame, text="2(택시/영업)", bg="skyblue", command=self.two_button)
        self.twoBtn.place(x=80, y=445, height=30, width=85)
        self.threeBtn = Button(self.frame, text="3(버스)", bg="skyblue", command=self.three_button)
        self.threeBtn.place(x=175, y=445, height=30, width=80)
        self.fourBtn = Button(self.frame, text="4(구)", bg="skyblue", command=self.four_button)
        self.fourBtn.place(x=265, y=445, height=30, width=65)
        self.fiveBtn = Button(self.frame, text="5(신)", bg="skyblue", command=self.five_button)
        self.fiveBtn.place(x=340, y=445, height=30, width=65)
        self.sixBtn = Button(self.frame, text="6(건설기계)", bg="skyblue", command=self.six_button)
        self.sixBtn.place(x=415, y=445, height=30, width=85)
        self.sevenBtn = Button(self.frame, text="7(긴외교)", bg="skyblue", command=self.seven_button)
        self.sevenBtn.place(x=510, y=445, height=30, width=75)
        self.eightBtn = Button(self.frame, text="8(짧은외교)", bg="skyblue", command=self.eight_button)
        self.eightBtn.place(x=595, y=445, height=30, width=85)
        self.nineBtn = Button(self.frame, text="9(임시)", bg="skyblue", command=self.nine_button)
        self.nineBtn.place(x=690, y=445, height=30, width=65)
        self.tenBtn = Button(self.frame, text="q(모름)", bg="skyblue", command=self.ten_button)
        self.tenBtn.place(x=765, y=445, height=30, width=65)

        # correct/incorrect button
        self.corBtn = Button(self.frame, text="O(맞음)", command=self.correct_button)
        self.corBtn.place(x=160, y=480, height=60, width=60)
        self.incorBtn = Button(self.frame, text="X(틀림)", command=self.incorrect_button)
        self.incorBtn.place(x=230, y=480, height=60, width=60)
        self.QBtn = Button(self.frame, text="?(오검지)", command=self.question_button)
        self.QBtn.place(x=300, y=480, height=60, width=60)
        self.XBtn = Button(self.frame, text="XX(인식불가)", command=self.X_button)
        self.XBtn.place(x=370, y=480, height=60, width=80)
        self.specialBtn = Button(self.frame, text="*(특수차량)", command=self.special_button)
        self.specialBtn.place(x=460, y=480, height=60, width=80)

        # next/prev button
        self.nextBtn = Button(self.frame, text="Next", font=('times new roman', 12, 'bold'), command=self.next_button)
        self.nextBtn.place(x=550, y=480, height=60, width=60)
        self.prevBtn = Button(self.frame, text="Prev", font=('times new roman', 12, 'bold'), command=self.prev_button)
        self.prevBtn.place(x=90, y=480, height=60, width=60)

        # save plt type image button
        self.saveBtn = Button(self.frame, text="Save Image", bg="skyblue", font=('times new roman', 12, 'bold'),
                              command=self.save_img)
        self.saveBtn.place(x=620, y=480, height=60, width=120)

        # go entry/button
        self.pageSrc = IntVar()
        self.go_entry = Entry(self.frame, textvariable=self.pageSrc)
        self.go_entry.place(x=750, y=500, height=20, width=30)
        self.go_button = Button(self.frame, text="go", command=self.go_specific_page)
        self.go_button.place(x=795, y=500, height=20, width=30)

        # detective image
        # self.mainPanel_5.create_image(125, 110, image=self.det_img)

        # keyboard manipulate
        self.parent.bind("z", lambda e: self.correct_button())
        self.parent.bind("x", lambda e: self.incorrect_button())
        self.parent.bind("c", lambda e: self.question_button())
        self.parent.bind("v", lambda e: self.X_button())
        self.parent.bind("b", lambda e: self.special_button())
        self.parent.bind("<Right>", lambda e: self.next_button())
        self.parent.bind("<Left>", lambda e: self.prev_button())
        self.parent.bind("q", lambda e: self.one_button())
        self.parent.bind("w", lambda e: self.two_button())
        self.parent.bind("e", lambda e: self.three_button())
        self.parent.bind("r", lambda e: self.four_button())
        self.parent.bind("t", lambda e: self.five_button())
        self.parent.bind("y", lambda e: self.six_button())
        self.parent.bind("u", lambda e: self.seven_button())
        self.parent.bind("i", lambda e: self.eight_button())
        self.parent.bind("o", lambda e: self.nine_button())
        self.parent.bind("p", lambda e: self.ten_button())
        self.parent.bind("<Control-s>", lambda e: self.save_img())

    def popup_for_labeling(self):
        self.Labeling = StringVar()
        self.labeling_popup = Tk()
        self.labeling_popup.title("Labeling Y/N")
        self.labeling_popup.geometry('{}x{}+{}+{}'.format(200, 120, 300, 300))
        self.label = Label(self.labeling_popup, text="Labeling?(Y/N)")
        self.label.grid(row=1, column=1, columnspan=2)
        self.entry = Entry(self.labeling_popup, textvariable=self.Labeling)
        self.entry.focus_set()
        self.entry.grid(row=2, column=1, pady=10, columnspan=2)
        self.submitBtn = Button(self.labeling_popup, text="Submit", command=self.cleanup_for_labeling)
        self.submitBtn.grid(row=3, column=1, pady=10)

    def popup_for_answer(self):
        self.answer = StringVar()
        self.popup = Tk()
        self.popup.title("Enter Right Answer")
        self.popup.bind('<Return>', lambda e: self.cleanup())
        # window_position(self.popup, width=200, height=120)
        self.popup.geometry('{}x{}+{}+{}'.format(200, 120, 300, 300))
        self.label = Label(self.popup, text="Enter Right Answer")
        self.label.grid(row=1, column=1, columnspan=2)
        self.entry = Entry(self.popup, textvariable=self.answer)
        self.entry.focus_set()
        self.entry.grid(row=2, column=1, pady=10, columnspan=2)
        self.submitBtn = Button(self.popup, text="Submit", command=self.cleanup)
        self.submitBtn.grid(row=3, column=1, pady=10)
        self.cancelBtn = Button(self.popup, text="Cancel", command=self.cancel)
        self.cancelBtn.grid(row=3, column=2, pady=10)

    def cleanup(self):
        self.answer_value = self.entry.get()
        self.answer_dict[self.cur] = self.answer_value
        self.popup.destroy()
        self.go_next()

    def cleanup_for_labeling(self):
        self.Labeling_result = self.entry.get()
        self.labeling_popup.destroy()

    def cancel(self):
        self.popup.destroy()
        self.incorBtn['state'] = NORMAL

    def select_input_dir(self):
        path = filedialog.askdirectory(title="Select image source folder", initialdir='./')
        self.source_dir.set(path)
        return

    def get_json_img_list(self):
        self.json_list = []
        self.img_list = []
        self.input_dir = self.source_dir.get()

        dir_list = os.listdir(self.input_dir)
        for dirs in dir_list:
            if os.path.isdir(os.path.join(self.input_dir, dirs)):
                files = os.listdir(os.path.join(self.input_dir, dirs))
                for file in files:
                    fname, ext = os.path.splitext(os.path.join(self.input_dir, dirs, file))
                    if ext == ".json":
                        self.json_list.append(os.path.join(self.input_dir, dirs, file))
                    elif ext == ".jpg":
                        self.img_list.append(os.path.join(self.input_dir, dirs, file))

            else:
                fname, ext = os.path.splitext(os.path.join(self.input_dir, dirs))
                if ext == ".json":
                    self.json_list.append(os.path.join(self.input_dir, dirs))
                elif ext == ".jpg":
                    self.img_list.append(os.path.join(self.input_dir, dirs))

    def get_common_name_list(self):
        self.cname_list = []
        for json_name in self.json_list:
            json_name = json_name.split('/')[-1]
            json_name_split = json_name.split('_')
            date, time = json_name_split[2], json_name_split[3]
            file_id = date + '_' + time
            if file_id not in self.cname_list:
                self.cname_list.append(file_id)

    def get_camera_info_from_json(self, file_id, postfix):
        for json_path in self.json_list:
            base_name = os.path.basename(json_path)
            last_number_str = os.path.splitext(base_name)[0][-1]
            if file_id in json_path and str(postfix) == last_number_str:
                with open(json_path, encoding='utf8') as json_file:
                    json_data = json.load(json_file)
                    for key in self.json_info_dict[str(postfix)]:
                        self.camera_info[key] = json_data[key]

    def load_img(self, uri, resize=None):
        local_uri = '/'.join([self.input_dir, uri.split('/')[-2], uri.split('/')[-1]])
        try:
            img = PIL.Image.open(local_uri)
            resizeimg = img.resize(resize)
            self.tkimg = PIL.ImageTk.PhotoImage(resizeimg)
            return self.img, self.tkimg

        except FileNotFoundError:
            # messagebox.showwarning("Warning", message='{} 파일이 존재하지 않습니다.'.format(local_uri))
            pass

    def load_img_keep_ratio(self, uri, width):
        local_uri = '/'.join([self.input_dir, uri.split('/')[-2], uri.split('/')[-1]])
        try:
            img = PIL.Image.open(local_uri)
            wpercent = (width/float(img.size[0]))
            hsize = int((float(img.size[1])*float(wpercent)))
            resizeimg = img.resize((width, hsize), PIL.Image.ANTIALIAS)
            self.tkimg = PIL.ImageTk.PhotoImage(resizeimg)
            return self.img, self.tkimg

        except FileNotFoundError:
            pass

    def show_img(self, camera_img=None, car_img=None, plt_img=None):

        self.mainPanel_1.create_image(250, 200, image=camera_img)
        self.mainPanel_2.create_image(150, 150, image=car_img)
        self.mainPanel_3.create_image(150, 50, image=plt_img)

    def show_text(self, camera_ID, Date, Time, type, color, dir, side, lane, Plate_Number, confidence, Now, Plt_type,
                  Answer, Edit):

        self.mainPanel_4.create_text(10, 40, font=("times new roman", 12, "bold"), fill="blue", text="camera ID",
                                     anchor=W)
        self.mainPanel_4.create_text(160, 40, font=("times new roman", 12), fill="black", text=camera_ID, anchor=W)
        self.mainPanel_4.create_text(10, 60, font=("times new roman", 12, "bold"), fill="blue", text="Date", anchor=W)
        self.mainPanel_4.create_text(160, 60, font=("times new roman", 12), fill="black", text=Date, anchor=W)
        self.mainPanel_4.create_text(10, 80, font=("times new roman", 12, "bold"), fill="blue", text="Time", anchor=W)
        self.mainPanel_4.create_text(160, 80, font=("times new roman", 12), fill="black", text=Time, anchor=W)
        self.mainPanel_4.create_text(10, 100, font=("times new roman", 12, "bold"), fill="blue", text="Type", anchor=W)
        self.mainPanel_4.create_text(160, 100, font=("times new roman", 12), fill="black", text=type, anchor=W)
        self.mainPanel_4.create_text(10, 120, font=("times new roman", 12, "bold"), fill="blue", text="Color", anchor=W)
        self.mainPanel_4.create_text(160, 120, font=("times new roman", 12), fill="black", text=color, anchor=W)
        self.mainPanel_4.create_text(10, 140, font=("times new roman", 12, "bold"), fill="blue", text="Direction",
                                     anchor=W)
        self.mainPanel_4.create_text(160, 140, font=("times new roman", 12), fill="black", text=dir, anchor=W)
        self.mainPanel_4.create_text(10, 160, font=("times new roman", 12, "bold"), fill="blue", text="Side", anchor=W)
        self.mainPanel_4.create_text(160, 160, font=("times new roman", 12), fill="black", text=side, anchor=W)
        self.mainPanel_4.create_text(10, 180, font=("times new roman", 12, "bold"), fill="blue", text="Lane", anchor=W)
        self.mainPanel_4.create_text(160, 180, font=("times new roman", 12), fill="black", text=lane, anchor=W)
        self.mainPanel_4.create_text(10, 300, font=("times new roman", 12, "bold"), fill="red", text="Plate Number",
                                     anchor=W)
        self.mainPanel_5.create_text(20, 40, font=("times new roman", 27, "bold"), fill="black", text=Plate_Number, anchor=W)
        self.mainPanel_4.create_text(10, 200, font=("times new roman", 12, "bold"), fill="blue", text="Confidence",
                                     anchor=W)
        self.mainPanel_4.create_text(160, 200, font=("times new roman", 12), fill="black", text=confidence, anchor=W)
        self.mainPanel_4.create_text(10, 220, font=("times new roman", 12, "bold"), fill="blue", text="Now / Total",
                                     anchor=W)
        self.mainPanel_4.create_text(160, 220, font=("times new roman", 12), fill="black", text=Now, anchor=W)
        self.mainPanel_4.create_text(10, 240, font=("times new roman", 12, "bold"), fill="blue", text="Plate Type",
                                     anchor=W)
        self.mainPanel_4.create_text(160, 240, font=("times new roman", 12), fill="black", text=Plt_type, anchor=W)
        self.mainPanel_4.create_text(10, 260, font=("times new roman", 12, "bold"), fill="blue", text="Answer",
                                     anchor=W)
        self.mainPanel_4.create_text(160, 260, font=("times new roman", 12), fill="black", text=Answer, anchor=W)
        self.mainPanel_4.create_text(10, 280, font=("times new roman", 12, "bold"), fill="blue", text="Edit",
                                     anchor=W)
        self.mainPanel_4.create_text(160, 280, font=("times new roman", 12), fill="black", text=Edit, anchor=W)

    def load_all(self):
        self.Labeling = self.popup_for_labeling()
        self.input_dir = ''
        self.json_list = []
        self.img_list = []
        self.cname_list = []
        self.camera_info = {}
        self.img = None
        self.camera_img = None
        self.car_img = None
        self.plt_img = None
        self.tkimg = None
        self.cur = 0
        self.plt_type_dict = {}
        self.result = {}
        self.result_merged_dict = {}
        self.mainPanel_1.delete('all')
        self.mainPanel_2.delete('all')
        self.mainPanel_3.delete('all')
        self.mainPanel_4.delete('all')
        self.mainPanel_5.delete('all')

        self.get_json_img_list()
        self.get_common_name_list()
        for idx, file_id in enumerate(self.cname_list):
            if idx == 0:
                # self.get_camera_info_from_json(file_id, postfix=0)
                self.get_camera_info_from_json(file_id, postfix=1)
                # try:
                #     self.img, self.camera_img = self.load_img(self.camera_info[camera_uri], resize=(500, 400))
                # except TypeError:
                #     self.camera_img = None
                try:
                    self.img, self.car_img = self.load_img(self.camera_info[veh_uri], resize=(300, 300))
                except TypeError:
                    self.car_img = None
                try:
                    # self.img, self.plt_img = self.load_img(self.camera_info[plate_uri], resize=(300, 100))
                    self.img, self.plt_img = self.load_img_keep_ratio(self.camera_info[plate_uri], width=300)
                except TypeError:
                    self.plt_img = None

                cam_id = self.camera_info[veh_uri].split('_')[-4]
                cam_date = self.camera_info[veh_uri].split('_')[-3]
                cam_time = self.camera_info[veh_uri].split('_')[-2]
                self.show_img(self.camera_img, self.car_img, self.plt_img)
                self.show_text(cam_id, cam_date, cam_time,
                               VehicleType(self.camera_info[veh_type]).name,
                               VehicleColor(self.camera_info[veh_color]).name,
                               VehicleDirection(self.camera_info[veh_dir]).name,
                               VehicleSide(self.camera_info[veh_side]).name,
                               self.camera_info[lane_info], self.camera_info[plate_num],
                               '{:4.2f}%'.format(self.camera_info[confidence]),
                               '{}/{}'.format(self.cur + 1, len(self.cname_list)), '', '', '')

    def one_button(self):
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL
        self.oneBtn['state'] = DISABLED

        self.plt_type_dict[self.cur] = ['1', self.camera_info[plate_uri], self.camera_info[plate_num]]

    def two_button(self):
        self.oneBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL
        self.twoBtn['state'] = DISABLED

        self.plt_type_dict[self.cur] = ['2', self.camera_info[plate_uri], self.camera_info[plate_num]]

    def three_button(self):
        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL
        self.threeBtn['state'] = DISABLED

        self.plt_type_dict[self.cur] = ['3', self.camera_info[plate_uri], self.camera_info[plate_num]]

    def four_button(self):
        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL
        self.fourBtn['state'] = DISABLED

        self.plt_type_dict[self.cur] = ['4', self.camera_info[plate_uri], self.camera_info[plate_num]]

    def five_button(self):
        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL
        self.fiveBtn['state'] = DISABLED

        self.plt_type_dict[self.cur] = ['5', self.camera_info[plate_uri], self.camera_info[plate_num]]

    def six_button(self):
        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL
        self.sixBtn['state'] = DISABLED

        self.plt_type_dict[self.cur] = ['6', self.camera_info[plate_uri], self.camera_info[plate_num]]

    def seven_button(self):
        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL
        self.sevenBtn['state'] = DISABLED

        self.plt_type_dict[self.cur] = ['7', self.camera_info[plate_uri], self.camera_info[plate_num]]

    def eight_button(self):
        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL
        self.eightBtn['state'] = DISABLED

        self.plt_type_dict[self.cur] = ['8', self.camera_info[plate_uri], self.camera_info[plate_num]]

    def nine_button(self):
        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL
        self.nineBtn['state'] = DISABLED

        self.plt_type_dict[self.cur] = ['9', self.camera_info[plate_uri], self.camera_info[plate_num]]

    def ten_button(self):
        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = DISABLED

        self.plt_type_dict[self.cur] = ['q', self.camera_info[plate_uri], self.camera_info[plate_num]]

    def correct_button(self):
        if self.incorBtn['state'] == DISABLED or self.QBtn['state'] == DISABLED:
            messagebox.showwarning("Warning", message='한가지만 선택하세요.')
        else:
            self.corBtn['state'] = DISABLED
            self.go_next()

        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL

    def incorrect_button(self):
        if self.incorBtn['state'] == DISABLED or self.QBtn['state'] == DISABLED:
            messagebox.showwarning("Warning", message='한가지만 선택하세요.')
        else:
            self.incorBtn['state'] = DISABLED
            if self.Labeling_result == "Y" or self.Labeling_result == "y":
                self.popup_for_answer()
            else:
                self.go_next()

        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL

    def question_button(self):
        self.QBtn['state'] = DISABLED
        self.go_next()

        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL

    def X_button(self):
        self.XBtn['state'] = DISABLED
        self.go_next()

        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL

    def special_button(self):
        self.specialBtn['state'] = DISABLED
        self.go_next()

        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL

    def next_button(self):

        if self.cur >= len(self.cname_list) - 1:
            messagebox.showwarning("Warning", message="마지막 사진입니다.")
        # increase
        else:
            self.cur += 1
            # if self.cur not in self.result.keys():
            #     self.result[self.cur] = None
            # load next
            for idx, file_id in enumerate(self.cname_list):
                if idx == self.cur:
                    self.mainPanel_1.delete('all')
                    self.mainPanel_2.delete('all')
                    self.mainPanel_3.delete('all')
                    self.mainPanel_4.delete('all')
                    self.mainPanel_5.delete('all')
                    # self.get_camera_info_from_json(file_id, postfix=0)
                    self.get_camera_info_from_json(file_id, postfix=1)
                    # try:
                    #     self.img, self.camera_img = self.load_img(self.camera_info[camera_uri], resize=(500, 400))
                    # except TypeError:
                    #     self.camera_img = None
                    try:
                        self.img, self.car_img = self.load_img(self.camera_info[veh_uri], resize=(300, 300))
                    except TypeError:
                        self.car_img = None
                    try:
                        # self.img, self.plt_img = self.load_img(self.camera_info[plate_uri], resize=(300, 100))
                        self.img, self.plt_img = self.load_img_keep_ratio(self.camera_info[plate_uri], width=300)
                    except TypeError:
                        self.plt_img = None
                    try:
                        Plt_type = self.plt_type_dict.get(self.cur)[0]
                    except TypeError:
                        Plt_type = ''
                    try:
                        Answer = self.result.get(self.cur)[-1]
                    except TypeError:
                        Answer = ''
                    try:
                        Edit = self.answer_dict[self.cur]
                    except KeyError:
                        Edit = ''

                    cam_id = self.camera_info[veh_uri].split('_')[-4]
                    cam_date = self.camera_info[veh_uri].split('_')[-3]
                    cam_time = self.camera_info[veh_uri].split('_')[-2]
                    self.show_img(self.camera_img, self.car_img, self.plt_img)
                    self.show_text(cam_id, cam_date, cam_time,
                                   VehicleType(self.camera_info[veh_type]).name,
                                   VehicleColor(self.camera_info[veh_color]).name,
                                   VehicleDirection(self.camera_info[veh_dir]).name,
                                   VehicleSide(self.camera_info[veh_side]).name,
                                   self.camera_info[lane_info], self.camera_info[plate_num],
                                   '{:4.2f}%'.format(self.camera_info[confidence]),
                                   '{}/{}'.format(self.cur + 1, len(self.cname_list)),
                                   Plt_type=Plt_type, Answer=Answer, Edit=Edit)

                    self.corBtn['state'] = NORMAL
                    self.incorBtn['state'] = NORMAL
                    self.QBtn['state'] = NORMAL
                    break

        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL

    def go_specific_page(self):
        self.cur = self.pageSrc.get() -1
        if int(self.cur) >= len(self.cname_list):
            messagebox.showwarning("Warning", message="존재하지 않는 이미지입니다.")

        else:
            for idx, file_id in enumerate(self.cname_list):
                if idx == int(self.cur):
                    self.mainPanel_1.delete('all')
                    self.mainPanel_2.delete('all')
                    self.mainPanel_3.delete('all')
                    self.mainPanel_4.delete('all')
                    self.mainPanel_5.delete('all')
                    # self.get_camera_info_from_json(file_id, postfix=0)
                    self.get_camera_info_from_json(file_id, postfix=1)
                    # try:
                    #     self.img, self.camera_img = self.load_img(self.camera_info[camera_uri], resize=(500, 400))
                    # except TypeError:
                    #     self.camera_img = None
                    try:
                        self.img, self.car_img = self.load_img(self.camera_info[veh_uri], resize=(300, 300))
                    except TypeError:
                        self.car_img = None
                    try:
                        # self.img, self.plt_img = self.load_img(self.camera_info[plate_uri], resize=(300, 100))
                        self.img, self.plt_img = self.load_img_keep_ratio(self.camera_info[plate_uri], width=300)
                    except TypeError:
                        self.plt_img = None
                    try:
                        Plt_type = self.plt_type_dict.get(self.cur)[0]
                    except TypeError:
                        Plt_type = ''
                    try:
                        Answer = self.result.get(self.cur)[-1]
                    except TypeError:
                        Answer = ''
                    try:
                        Edit = self.answer_dict[self.cur]
                    except KeyError:
                        Edit = ''

                    cam_id = self.camera_info[veh_uri].split('_')[-4]
                    cam_date = self.camera_info[veh_uri].split('_')[-3]
                    cam_time = self.camera_info[veh_uri].split('_')[-2]
                    self.show_img(self.camera_img, self.car_img, self.plt_img)
                    self.show_text(cam_id, cam_date, cam_time,
                                   VehicleType(self.camera_info[veh_type]).name,
                                   VehicleColor(self.camera_info[veh_color]).name,
                                   VehicleDirection(self.camera_info[veh_dir]).name,
                                   VehicleSide(self.camera_info[veh_side]).name,
                                   self.camera_info[lane_info], self.camera_info[plate_num],
                                   '{:4.2f}%'.format(self.camera_info[confidence]),
                                   '{}/{}'.format(self.cur + 1, len(self.cname_list)),
                                   Plt_type=Plt_type, Answer=Answer, Edit=Edit)

                    self.corBtn['state'] = NORMAL
                    self.incorBtn['state'] = NORMAL
                    self.QBtn['state'] = NORMAL
                    break

        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL
        self.go_button.focus_set()

    def go_next(self):

        if self.corBtn['state'] == DISABLED and self.incorBtn['state'] == NORMAL and self.QBtn['state'] == NORMAL and \
                self.XBtn['state'] == NORMAL and self.specialBtn['state'] == NORMAL:
            # append
            now_cur, now_camera_id, now_date, now_time, car_uri, plt_uri, now_type, now_color, now_dir, now_side, \
            now_lane, now_plate_num, now_conf, correct = self.cur, self.camera_info[veh_uri].split('_')[-4], \
                                                         self.camera_info[veh_uri].split('_')[-3], \
                                                         self.camera_info[veh_uri].split('_')[-2], \
                                                         self.camera_info[veh_uri], self.camera_info[plate_uri], \
                                                         VehicleType(self.camera_info[veh_type]).name, \
                                                         VehicleColor(self.camera_info[veh_color]).name, \
                                                         VehicleDirection(self.camera_info[veh_dir]).name, \
                                                         VehicleSide(self.camera_info[veh_side]).name, \
                                                         self.camera_info[lane_info], self.camera_info[plate_num], \
                                                         self.camera_info[confidence], "O"
            # ,self.camera_info[camera_uri]
            self.result[now_cur] = [now_cur, now_camera_id, now_date, now_time, car_uri, plt_uri, now_type,
                                    now_color, now_dir, now_side, now_lane, now_plate_num, now_conf, correct]
            # increase
            if self.cur < len(self.cname_list)-1:
                self.cur += 1
            # load next
            for idx, file_id in enumerate(self.cname_list):
                if idx == self.cur:
                    self.mainPanel_1.delete('all')
                    self.mainPanel_2.delete('all')
                    self.mainPanel_3.delete('all')
                    self.mainPanel_4.delete('all')
                    self.mainPanel_5.delete('all')
                    # self.get_camera_info_from_json(file_id, postfix=0)
                    self.get_camera_info_from_json(file_id, postfix=1)
                    # try:
                    #     self.img, self.camera_img = self.load_img(self.camera_info[camera_uri], resize=(500, 400))
                    # except TypeError:
                    #     self.camera_img = None
                    try:
                        self.img, self.car_img = self.load_img(self.camera_info[veh_uri], resize=(300, 300))
                    except TypeError:
                        self.car_img = None
                    try:
                        # self.img, self.plt_img = self.load_img(self.camera_info[plate_uri], resize=(300, 100))
                        self.img, self.plt_img = self.load_img_keep_ratio(self.camera_info[plate_uri], width=300)
                    except TypeError:
                        self.plt_img = None

                    self.show_img(self.camera_img, self.car_img, self.plt_img)

                    try:
                        Plt_type = self.plt_type_dict.get(self.cur)[0]
                    except TypeError:
                        Plt_type = ''
                    try:
                        Answer = self.result.get(self.cur)[-1]
                    except TypeError:
                        Answer = ''
                    try:
                        Edit = self.answer_dict[self.cur]
                    except KeyError:
                        Edit = ''

                    cam_id = self.camera_info[veh_uri].split('_')[-4]
                    cam_date = self.camera_info[veh_uri].split('_')[-3]
                    cam_time = self.camera_info[veh_uri].split('_')[-2]
                    self.show_text(cam_id, cam_date, cam_time,
                                   VehicleType(self.camera_info[veh_type]).name,
                                   VehicleColor(self.camera_info[veh_color]).name,
                                   VehicleDirection(self.camera_info[veh_dir]).name,
                                   VehicleSide(self.camera_info[veh_side]).name,
                                   self.camera_info[lane_info], self.camera_info[plate_num],
                                   '{:4.2f}%'.format(self.camera_info[confidence]),
                                   '{}/{}'.format(self.cur + 1, len(self.cname_list)),
                                   Plt_type=Plt_type, Answer=Answer, Edit=Edit)

                    self.corBtn['state'] = NORMAL
                    self.incorBtn['state'] = NORMAL
                    self.QBtn['state'] = NORMAL
                    self.XBtn['state'] = NORMAL
                    self.specialBtn['state'] = NORMAL
                    break

        if self.corBtn['state'] == NORMAL and self.incorBtn['state'] == DISABLED and self.QBtn['state'] == NORMAL and \
                self.XBtn['state'] == NORMAL and self.specialBtn['state'] == NORMAL:
            # append
            now_cur, now_camera_id, now_date, now_time, car_uri, plt_uri, now_type, now_color, now_dir, now_side, \
            now_lane, now_plate_num, now_conf, correct = self.cur, self.camera_info[veh_uri].split('_')[-4], \
                                                         self.camera_info[veh_uri].split('_')[-3], \
                                                         self.camera_info[veh_uri].split('_')[-2], \
                                                         self.camera_info[veh_uri], self.camera_info[plate_uri], \
                                                         VehicleType(self.camera_info[veh_type]).name, \
                                                         VehicleColor(self.camera_info[veh_color]).name, \
                                                         VehicleDirection(self.camera_info[veh_dir]).name, \
                                                         VehicleSide(self.camera_info[veh_side]).name, \
                                                         self.camera_info[lane_info], self.camera_info[plate_num], \
                                                         self.camera_info[confidence], "X"
            # ,self.camera_info[camera_uri]
            self.result[now_cur] = [now_cur, now_camera_id, now_date, now_time, car_uri, plt_uri, now_type,
                                    now_color, now_dir, now_side, now_lane, now_plate_num, now_conf, correct]
            # increase
            if self.cur < len(self.cname_list)-1:
                self.cur += 1
            # load next
            for idx, file_id in enumerate(self.cname_list):
                if idx == self.cur:
                    self.mainPanel_1.delete('all')
                    self.mainPanel_2.delete('all')
                    self.mainPanel_3.delete('all')
                    self.mainPanel_4.delete('all')
                    self.mainPanel_5.delete('all')
                    # self.get_camera_info_from_json(file_id, postfix=0)
                    self.get_camera_info_from_json(file_id, postfix=1)
                    # try:
                    #     self.img, self.camera_img = self.load_img(self.camera_info[camera_uri], resize=(500, 400))
                    # except TypeError:
                    #     self.camera_img = None
                    try:
                        self.img, self.car_img = self.load_img(self.camera_info[veh_uri], resize=(300, 300))
                    except TypeError:
                        self.car_img = None
                    try:
                        # self.img, self.plt_img = self.load_img(self.camera_info[plate_uri], resize=(300, 100))
                        self.img, self.plt_img = self.load_img_keep_ratio(self.camera_info[plate_uri], width=300)
                    except TypeError:
                        self.plt_img = None

                    self.show_img(self.camera_img, self.car_img, self.plt_img)

                    try:
                        Plt_type = self.plt_type_dict.get(self.cur)[0]
                    except TypeError:
                        Plt_type = ''
                    try:
                        Answer = self.result.get(self.cur)[-1]
                    except TypeError:
                        Answer = ''
                    try:
                        Edit = self.answer_dict[self.cur]
                    except KeyError:
                        Edit = ''

                    cam_id = self.camera_info[veh_uri].split('_')[-4]
                    cam_date = self.camera_info[veh_uri].split('_')[-3]
                    cam_time = self.camera_info[veh_uri].split('_')[-2]
                    self.show_text(cam_id, cam_date, cam_time,
                                   VehicleType(self.camera_info[veh_type]).name,
                                   VehicleColor(self.camera_info[veh_color]).name,
                                   VehicleDirection(self.camera_info[veh_dir]).name,
                                   VehicleSide(self.camera_info[veh_side]).name,
                                   self.camera_info[lane_info], self.camera_info[plate_num],
                                   '{:4.2f}%'.format(self.camera_info[confidence]),
                                   '{}/{}'.format(self.cur + 1, len(self.cname_list)),
                                   Plt_type=Plt_type, Answer=Answer, Edit=Edit)

                    self.corBtn['state'] = NORMAL
                    self.incorBtn['state'] = NORMAL
                    self.QBtn['state'] = NORMAL
                    self.XBtn['state'] = NORMAL
                    self.specialBtn['state'] = NORMAL
                    break

        if self.corBtn['state'] == NORMAL and self.incorBtn['state'] == NORMAL and self.QBtn['state'] == DISABLED and \
                self.XBtn['state'] == NORMAL and self.specialBtn['state'] == NORMAL:
            # append
            now_cur, now_camera_id, now_date, now_time, car_uri, plt_uri, now_type, now_color, now_dir, now_side, \
            now_lane, now_plate_num, now_conf, correct = self.cur, self.camera_info[veh_uri].split('_')[-4], \
                                                         self.camera_info[veh_uri].split('_')[-3], \
                                                         self.camera_info[veh_uri].split('_')[-2], \
                                                         self.camera_info[veh_uri], self.camera_info[plate_uri], \
                                                         VehicleType(self.camera_info[veh_type]).name, \
                                                         VehicleColor(self.camera_info[veh_color]).name, \
                                                         VehicleDirection(self.camera_info[veh_dir]).name, \
                                                         VehicleSide(self.camera_info[veh_side]).name, \
                                                         self.camera_info[lane_info], self.camera_info[plate_num], \
                                                         self.camera_info[confidence], "q"
            # ,self.camera_info[camera_uri]
            self.result[now_cur] = [now_cur, now_camera_id, now_date, now_time, car_uri, plt_uri, now_type,
                                    now_color, now_dir, now_side, now_lane, now_plate_num, now_conf, correct]
            # increase
            if self.cur < len(self.cname_list)-1:
                self.cur += 1
            # load next
            for idx, file_id in enumerate(self.cname_list):
                if idx == self.cur:
                    self.mainPanel_1.delete('all')
                    self.mainPanel_2.delete('all')
                    self.mainPanel_3.delete('all')
                    self.mainPanel_4.delete('all')
                    self.mainPanel_5.delete('all')
                    # self.get_camera_info_from_json(file_id, postfix=0)
                    self.get_camera_info_from_json(file_id, postfix=1)
                    # try:
                    #     self.img, self.camera_img = self.load_img(self.camera_info[camera_uri], resize=(500, 400))
                    # except TypeError:
                    #     self.camera_img = None
                    try:
                        self.img, self.car_img = self.load_img(self.camera_info[veh_uri], resize=(300, 300))
                    except TypeError:
                        self.car_img = None
                    try:
                        # self.img, self.plt_img = self.load_img(self.camera_info[plate_uri], resize=(300, 100))
                        self.img, self.plt_img = self.load_img_keep_ratio(self.camera_info[plate_uri], width=300)
                    except TypeError:
                        self.plt_img = None

                    self.show_img(self.camera_img, self.car_img, self.plt_img)

                    try:
                        Plt_type = self.plt_type_dict.get(self.cur)[0]
                    except TypeError:
                        Plt_type = ''
                    try:
                        Answer = self.result.get(self.cur)[-1]
                    except TypeError:
                        Answer = ''
                    try:
                        Edit = self.answer_dict[self.cur]
                    except KeyError:
                        Edit = ''

                    cam_id = self.camera_info[veh_uri].split('_')[-4]
                    cam_date = self.camera_info[veh_uri].split('_')[-3]
                    cam_time = self.camera_info[veh_uri].split('_')[-2]
                    self.show_text(cam_id, cam_date, cam_time,
                                   VehicleType(self.camera_info[veh_type]).name,
                                   VehicleColor(self.camera_info[veh_color]).name,
                                   VehicleDirection(self.camera_info[veh_dir]).name,
                                   VehicleSide(self.camera_info[veh_side]).name,
                                   self.camera_info[lane_info], self.camera_info[plate_num],
                                   '{:4.2f}%'.format(self.camera_info[confidence]),
                                   '{}/{}'.format(self.cur + 1, len(self.cname_list)),
                                   Plt_type=Plt_type, Answer=Answer, Edit=Edit)

                    self.corBtn['state'] = NORMAL
                    self.incorBtn['state'] = NORMAL
                    self.QBtn['state'] = NORMAL
                    self.XBtn['state'] = NORMAL
                    self.specialBtn['state'] = NORMAL
                    break

        if self.corBtn['state'] == NORMAL and self.incorBtn['state'] == NORMAL and self.QBtn['state'] == NORMAL and \
                self.XBtn['state'] == DISABLED and self.specialBtn['state'] == NORMAL:
            # append
            now_cur, now_camera_id, now_date, now_time, car_uri, plt_uri, now_type, now_color, now_dir, now_side, \
            now_lane, now_plate_num, now_conf, correct = self.cur, self.camera_info[veh_uri].split('_')[-4], \
                                                         self.camera_info[veh_uri].split('_')[-3], \
                                                         self.camera_info[veh_uri].split('_')[-2], \
                                                         self.camera_info[veh_uri], self.camera_info[plate_uri], \
                                                         VehicleType(self.camera_info[veh_type]).name, \
                                                         VehicleColor(self.camera_info[veh_color]).name, \
                                                         VehicleDirection(self.camera_info[veh_dir]).name, \
                                                         VehicleSide(self.camera_info[veh_side]).name, \
                                                         self.camera_info[lane_info], self.camera_info[plate_num], \
                                                         self.camera_info[confidence], "XX"
            # ,self.camera_info[camera_uri]
            self.result[now_cur] = [now_cur, now_camera_id, now_date, now_time, car_uri, plt_uri, now_type,
                                    now_color, now_dir, now_side, now_lane, now_plate_num, now_conf, correct]
            # increase
            if self.cur < len(self.cname_list) - 1:
                self.cur += 1
            # load next
            for idx, file_id in enumerate(self.cname_list):
                if idx == self.cur:
                    self.mainPanel_1.delete('all')
                    self.mainPanel_2.delete('all')
                    self.mainPanel_3.delete('all')
                    self.mainPanel_4.delete('all')
                    self.mainPanel_5.delete('all')
                    # self.get_camera_info_from_json(file_id, postfix=0)
                    self.get_camera_info_from_json(file_id, postfix=1)
                    # try:
                    #     self.img, self.camera_img = self.load_img(self.camera_info[camera_uri], resize=(500, 400))
                    # except TypeError:
                    #     self.camera_img = None
                    try:
                        self.img, self.car_img = self.load_img(self.camera_info[veh_uri], resize=(300, 300))
                    except TypeError:
                        self.car_img = None
                    try:
                        # self.img, self.plt_img = self.load_img(self.camera_info[plate_uri], resize=(300, 100))
                        self.img, self.plt_img = self.load_img_keep_ratio(self.camera_info[plate_uri], width=300)
                    except TypeError:
                        self.plt_img = None

                    self.show_img(self.camera_img, self.car_img, self.plt_img)

                    try:
                        Plt_type = self.plt_type_dict.get(self.cur)[0]
                    except TypeError:
                        Plt_type = ''
                    try:
                        Answer = self.result.get(self.cur)[-1]
                    except TypeError:
                        Answer = ''
                    try:
                        Edit = self.answer_dict[self.cur]
                    except KeyError:
                        Edit = ''

                    cam_id = self.camera_info[veh_uri].split('_')[-4]
                    cam_date = self.camera_info[veh_uri].split('_')[-3]
                    cam_time = self.camera_info[veh_uri].split('_')[-2]
                    self.show_text(cam_id, cam_date, cam_time,
                                   VehicleType(self.camera_info[veh_type]).name,
                                   VehicleColor(self.camera_info[veh_color]).name,
                                   VehicleDirection(self.camera_info[veh_dir]).name,
                                   VehicleSide(self.camera_info[veh_side]).name,
                                   self.camera_info[lane_info], self.camera_info[plate_num],
                                   '{:4.2f}%'.format(self.camera_info[confidence]),
                                   '{}/{}'.format(self.cur + 1, len(self.cname_list)),
                                   Plt_type=Plt_type, Answer=Answer, Edit=Edit)

                    self.corBtn['state'] = NORMAL
                    self.incorBtn['state'] = NORMAL
                    self.QBtn['state'] = NORMAL
                    self.XBtn['state'] = NORMAL
                    self.specialBtn['state'] = NORMAL
                    break

        if self.corBtn['state'] == NORMAL and self.incorBtn['state'] == NORMAL and self.QBtn['state'] == NORMAL and \
                self.XBtn['state'] == NORMAL and self.specialBtn['state'] == DISABLED:
            # append
            now_cur, now_camera_id, now_date, now_time, car_uri, plt_uri, now_type, now_color, now_dir, now_side, \
            now_lane, now_plate_num, now_conf, correct = self.cur, self.camera_info[veh_uri].split('_')[-4], \
                                                         self.camera_info[veh_uri].split('_')[-3], \
                                                         self.camera_info[veh_uri].split('_')[-2], \
                                                         self.camera_info[veh_uri], self.camera_info[plate_uri], \
                                                         VehicleType(self.camera_info[veh_type]).name, \
                                                         VehicleColor(self.camera_info[veh_color]).name, \
                                                         VehicleDirection(self.camera_info[veh_dir]).name, \
                                                         VehicleSide(self.camera_info[veh_side]).name, \
                                                         self.camera_info[lane_info], self.camera_info[plate_num], \
                                                         self.camera_info[confidence], "*"
            # ,self.camera_info[camera_uri]
            self.result[now_cur] = [now_cur, now_camera_id, now_date, now_time, car_uri, plt_uri, now_type,
                                    now_color, now_dir, now_side, now_lane, now_plate_num, now_conf, correct]
            # increase
            if self.cur < len(self.cname_list)-1:
                self.cur += 1
            # load next
            for idx, file_id in enumerate(self.cname_list):
                if idx == self.cur:
                    self.mainPanel_1.delete('all')
                    self.mainPanel_2.delete('all')
                    self.mainPanel_3.delete('all')
                    self.mainPanel_4.delete('all')
                    self.mainPanel_5.delete('all')
                    # self.get_camera_info_from_json(file_id, postfix=0)
                    self.get_camera_info_from_json(file_id, postfix=1)
                    # try:
                    #     self.img, self.camera_img = self.load_img(self.camera_info[camera_uri], resize=(500, 400))
                    # except TypeError:
                    #     self.camera_img = None
                    try:
                        self.img, self.car_img = self.load_img(self.camera_info[veh_uri], resize=(300, 300))
                    except TypeError:
                        self.car_img = None
                    try:
                        # self.img, self.plt_img = self.load_img(self.camera_info[plate_uri], resize=(300, 100))
                        self.img, self.plt_img = self.load_img_keep_ratio(self.camera_info[plate_uri], width=300)
                    except TypeError:
                        self.plt_img = None

                    self.show_img(self.camera_img, self.car_img, self.plt_img)

                    try:
                        Plt_type = self.plt_type_dict.get(self.cur)[0]
                    except TypeError:
                        Plt_type = ''
                    try:
                        Answer = self.result.get(self.cur)[-1]
                    except TypeError:
                        Answer = ''
                    try:
                        Edit = self.answer_dict[self.cur]
                    except KeyError:
                        Edit = ''

                    cam_id = self.camera_info[veh_uri].split('_')[-4]
                    cam_date = self.camera_info[veh_uri].split('_')[-3]
                    cam_time = self.camera_info[veh_uri].split('_')[-2]
                    self.show_text(cam_id, cam_date, cam_time,
                                   VehicleType(self.camera_info[veh_type]).name,
                                   VehicleColor(self.camera_info[veh_color]).name,
                                   VehicleDirection(self.camera_info[veh_dir]).name,
                                   VehicleSide(self.camera_info[veh_side]).name,
                                   self.camera_info[lane_info], self.camera_info[plate_num],
                                   '{:4.2f}%'.format(self.camera_info[confidence]),
                                   '{}/{}'.format(self.cur + 1, len(self.cname_list)),
                                   Plt_type=Plt_type, Answer=Answer, Edit=Edit)

                    self.corBtn['state'] = NORMAL
                    self.incorBtn['state'] = NORMAL
                    self.QBtn['state'] = NORMAL
                    self.XBtn['state'] = NORMAL
                    self.specialBtn['state'] = NORMAL
                    break

    def prev_button(self):

        if self.cur == 0:
            messagebox.showwarning("Warning", message="첫번째 사진입니다.")

        else:
            self.cur -= 1
            for idx, file_id in enumerate(self.cname_list):
                if idx == self.cur:
                    self.mainPanel_1.delete('all')
                    self.mainPanel_2.delete('all')
                    self.mainPanel_3.delete('all')
                    self.mainPanel_4.delete('all')
                    self.mainPanel_5.delete('all')
                    # self.get_camera_info_from_json(file_id, postfix=0)
                    self.get_camera_info_from_json(file_id, postfix=1)
                    # try:
                    #     self.img, self.camera_img = self.load_img(self.camera_info[camera_uri], resize=(500, 400))
                    # except TypeError:
                    #     self.camera_img = None
                    try:
                        self.img, self.car_img = self.load_img(self.camera_info[veh_uri], resize=(300, 300))
                    except TypeError:
                        self.car_img = None
                    try:
                        # self.img, self.plt_img = self.load_img(self.camera_info[plate_uri], resize=(300, 100))
                        self.img, self.plt_img = self.load_img_keep_ratio(self.camera_info[plate_uri], width=300)
                    except TypeError:
                        self.plt_img = None

                    self.show_img(self.camera_img, self.car_img, self.plt_img)

                    try:
                        Plt_type = self.plt_type_dict.get(self.cur)[0]
                    except TypeError:
                        Plt_type = ''
                    try:
                        Answer = self.result.get(self.cur)[-1]
                    except TypeError:
                        Answer = ''
                    try:
                        Edit = self.answer_dict[self.cur]
                    except KeyError:
                        Edit = ''

                    cam_id = self.camera_info[veh_uri].split('_')[-4]
                    cam_date = self.camera_info[veh_uri].split('_')[-3]
                    cam_time = self.camera_info[veh_uri].split('_')[-2]
                    self.show_text(cam_id, cam_date, cam_time,
                                   VehicleType(self.camera_info[veh_type]).name,
                                   VehicleColor(self.camera_info[veh_color]).name,
                                   VehicleDirection(self.camera_info[veh_dir]).name,
                                   VehicleSide(self.camera_info[veh_side]).name,
                                   self.camera_info[lane_info], self.camera_info[plate_num],
                                   '{:4.2f}%'.format(self.camera_info[confidence]),
                                   '{}/{}'.format(self.cur + 1, len(self.cname_list)),
                                   Plt_type=Plt_type, Answer=Answer, Edit=Edit)

                    self.corBtn['state'] = NORMAL
                    self.incorBtn['state'] = NORMAL
                    self.QBtn['state'] = NORMAL
                    break

        self.oneBtn['state'] = NORMAL
        self.twoBtn['state'] = NORMAL
        self.threeBtn['state'] = NORMAL
        self.fourBtn['state'] = NORMAL
        self.fiveBtn['state'] = NORMAL
        self.sixBtn['state'] = NORMAL
        self.sevenBtn['state'] = NORMAL
        self.eightBtn['state'] = NORMAL
        self.nineBtn['state'] = NORMAL
        self.tenBtn['state'] = NORMAL

    def export_button(self):
        self.export_excel(os.path.join(self.input_dir,
                                       '{}_{}_result.xlsx'.format(self.camera_info[veh_uri].split('_')[-4],
                                                                  self.camera_info[veh_uri].split('_')[-3])))

    def export_excel(self, title):

        cor_num, incor_num, quest_num, impossible_num, special_num, accuracy = get_accuracy(self.result)
        acc_result_header = ['O 개수', 'X 개수', '? 개수', 'XX개수', '특수차량 수', '정확도']
        acc_result = [cor_num, incor_num, quest_num, impossible_num, special_num, accuracy]

        self.result_merged_dict = merge_dicts(self.result, self.plt_type_dict)

        if os.path.isfile(title):
            wb = load_workbook(title)
            ws = wb.active

            # no_answer_cnt = 0
            # no_answer_list = []
            # no_type_cnt = 0
            # no_type_list = []
            #
            # for i in range(len(self.cname_list)):
            #     try:
            #         self.result_merged_dict[i]
            #     except KeyError:
            #         messagebox.showwarning("Warning", text='정답과 번호판 타입이 모두 선택되지 않은 이미지가 있습니다.')
            #
            # for key, value in sorted(self.result_merged_dict.items()):
            #     if len(value) == 0:
            #         no_answer_cnt += 1
            #         no_answer_list.append(int(key+1))
            #         no_type_cnt += 1
            #         no_type_list.append(int(key+1))
            #
            #     elif len(value) <= 3:
            #         no_answer_cnt += 1
            #         no_answer_list.append(int(key+1))
            #
            #     elif len(value) <= 15:
            #         no_type_cnt += 1
            #         no_type_list.append(int(key+1))
            #
            # if no_type_cnt != 0 or no_answer_cnt != 0:
            #     messagebox.showwarning("Warning",
            #                            message='{}개 정답이 선택되지 않았습니다.({})\n{}개 번호판 타입이 선택되지 않았습니다.({})'.format
            #                            (no_answer_cnt, no_answer_list, no_type_cnt, no_type_list))
            #
            # else:
            for key, value in sorted(self.result_merged_dict.items()):
                ws.append(value)

            ws.append(acc_result_header)
            ws.append(acc_result)

            wb.save(title)
            messagebox.showinfo("Info", message="엑셀 파일 저장이 완료되었습니다.")

        else:
            wb = Workbook()
            ws = wb.active

            header_list = ['NO', 'ID', '날짜', '시간', '자동차 uri', '번호판 uri', '차량 종류', '차량 색상',
                           '차량 방향', '전후면 구분', '차선 정보', '번호판', '신뢰도', '검지 정답 여부', '번호판 종류']
            for i in range(len(header_list)):
                ws.cell(row=1, column=i + 1).fill = PatternFill(patternType='solid', fgColor=Color('FFC000'))
                ws.cell(row=1, column=i + 1).value = header_list[i]

            # no_answer_cnt = 0
            # no_answer_list = []
            # no_type_cnt = 0
            # no_type_list = []
            # for key, value in sorted(self.result_merged_dict.items()):
            #     if len(value) <= 3:
            #         no_answer_cnt += 1
            #         no_answer_list.append(int(key+1))
            #
            #     elif len(value) <= 15:
            #         no_type_cnt += 1
            #         no_type_list.append(int(key+1))
            #
            # if no_type_cnt != 0 or no_answer_cnt != 0:
            #     messagebox.showwarning("Warning",
            #                            message='{}개 정답이 선택되지 않았습니다.{}\n{}개 이미지의 번호판 타입이 선택되지 않았습니다.{}'.format
            #                            (no_answer_cnt, no_answer_list, no_type_cnt, no_type_list))
            #
            # else:
            for key, value in sorted(self.result_merged_dict.items()):
                ws.append(value)

            ws.append(acc_result_header)
            ws.append(acc_result)

            wb.save(title)
            messagebox.showinfo("Info", message="엑셀 파일 저장이 완료되었습니다.")

    def save_img(self):
        for key, value in self.plt_type_dict.items():
            out_dir_path = os.path.join(self.input_dir, "_".join(["type", value[0]]))
            in_dir_path = os.path.join(self.input_dir, "_".join(["type", value[0]]), self.result[key][-1])
            img_name = '/'.join([value[1].split('/')[-2], value[1].split('/')[-1]])
            if not os.path.exists(out_dir_path):
                os.mkdir(out_dir_path)
            if not os.path.exists(in_dir_path):
                os.mkdir(in_dir_path)
            tmp_img = PIL.Image.open(os.path.join(self.input_dir, img_name))
            try:
                self.answer_dict[key]
                tmp_img.save(''.join([os.path.join(in_dir_path, self.answer_dict[key]), '.jpg']))
            except Exception as e1:
                print("exception1:{}".format(e1))
                print(traceback.format_exc())
                try:
                    if value[2] == "":
                        tmp_img.save(os.path.join(in_dir_path, value[1].split('/')[-1]))
                    else:
                        tmp_img.save(''.join([os.path.join(in_dir_path, value[2].replace('*','-')), '.jpg']))
                except Exception as e2:
                    print("exception2:{}".format(e2))
                    print(traceback.format_exc())

        messagebox.showinfo("Info", message="번호판 타입 이미지 저장이 완료되었습니다.")


def main():
    root = Tk()
    tool = DetectCheck(root)
    window_position(root, width=1100, height=550)
    root.resizable(width=True, height=True)
    root.mainloop()


if __name__ == '__main__':
    main()
